from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from inventory.forms import UserRegistry, ProductForm, OrderForm, StoreForm, SupplierForm
from inventory.models import Product, Order, Store, dispatched, Supplier
from django.contrib.auth.views import LoginView

class CustomLoginView(LoginView):
    def get_success_url(self):
        if self.request.user.is_superuser:
            return 'dash'  # Redirect to the superuser dashboard
        else:
            return 'orders'  # Redirect to a generic dashboard for other roles

@login_required
def index(request):
    orders_user = Order.objects.all()
    users = User.objects.all()[:2]
    orders_adm = Order.objects.all()[:2]
    products = Product.objects.all()[:2]
    stores = Store.objects.all()[:2]
    dispatchs = dispatched.objects.all()[:2]
    reg_users = len(User.objects.all())
    all_prods = len(Product.objects.all())
    all_orders = len(Order.objects.all())
    all_stores = len(Store.objects.all())
    all_dispatchs = len(dispatched.objects.all())
    context = {
        "title": "Home",
        "orders": orders_user,
        "orders_adm": orders_adm,
        "users": users,
        "products": products,
        "stores": stores,
        "dispatchs": dispatchs,
        "count_users": reg_users,
        "count_products": all_prods,
        "count_orders": all_orders,
        "count_stores": all_stores,
        "count_dispatchs": all_dispatchs,

    }
    return render(request, 'inventory/index.html', context)

@login_required
def products(request):
    # language = 'am'
    # trans = translate(language)
    products = Product.objects.all()
    
    if request.method == 'POST':
        form = ProductForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('products')
    else:
        form = ProductForm()
    
    context = {
        "title": "Products",
        "products": products,
        "form": form,
        # "trans": trans
    }
    
    return render(request, 'inventory/products.html', context)


from django.shortcuts import render, redirect
from django.contrib import messages
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.conf import settings

from .models import Order, Product
from .forms import OrderForm
from django.contrib import messages

def orders(request):
    orders = Order.objects.all()

    if request.method == 'POST':
        form = OrderForm(request.POST)
        if form.is_valid():
            order = form.save(commit=False)
            product = order.product
            order.created_by = request.user

            if order.order_quantity > product.quantity:
                messages.error(request, "Not enough quantity available.")
            else:
                # Subtract order_quantity from the original quantity
                product.quantity -= order.order_quantity
                product.save()

                if product.quantity <= 3:
                    # Display alert message to superusers
                    alert_message = f"The remaining quantity of {product.name} is {product.quantity}."
                    for user in User.objects.filter(is_superuser=True):
                        messages.warning(request, alert_message, extra_tags='alert')

                        # Send email to superusers
                        subject = 'Low Inventory Alert'
                        context = {'product_name': product.name, 'product_quantity': product.quantity, 'message': alert_message}
                        html_message = render_to_string('inventory/low_inventory_email.html', context)
                        plain_message = strip_tags(html_message)
                        from_email = 'empireboss7887@gmail.com'
                        to_email = [user.email]
                        send_mail(subject, plain_message, from_email, to_email, html_message=html_message)

                order.save()
                return redirect('orders')
    else:
        form = OrderForm()

    context = {
        "title": "Orders",
        "orders": orders,
        "form": form
    }
    return render(request, 'inventory/orders.html', context)





@login_required
def users(request):
    users = User.objects.all()
    context = {
        "title": "Users",
        "users": users
    }
    return render(request, 'inventory/users.html', context)

@login_required
def user(request):
    context = {
        "profile": "User Profile"
    }
    return render(request, 'inventory/user.html', context)

def register(request):
    if request.method == 'POST':
        form = UserRegistry(request.POST)
        if form.is_valid():
            form.save()
            return redirect('login')
    else:
        form = UserRegistry()
    context = {
        "register": "Register",
        "form": form
    }
    return render(request, 'inventory/register.html', context)


from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
@login_required

def export_receipt(request, order_id):
    # Retrieve the order based on the provided order_id
    order = get_object_or_404(Order, id=order_id)

    # Create a new Excel workbook and select the active worksheet
    workbook = Workbook()
    worksheet = workbook.active

    # Set column widths for the first 6 columns
    column_widths = [20, 30, 15, 20, 20, 10]  # Adjust the widths as per your requirements
    for col_num, width in enumerate(column_widths, 1):
        col_letter = get_column_letter(col_num)
        worksheet.column_dimensions[col_letter].width = width

    # Write the title
    title = "INVENTORY MANAGEMENT SYSTEM"
    title_cell = worksheet.cell(row=1, column=1, value=title)
    title_cell.alignment = Alignment(horizontal='center')
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    # Set column headers
    headers = ['Order ID', 'Product Name', 'Quantity', 'Price', 'Created By', 'Date']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = worksheet.cell(row=2, column=col_num, value=header)
        cell.alignment = Alignment(horizontal='center')

    # Write the receipt details to the worksheet
    row_num = 3
    worksheet.cell(row=row_num, column=1, value=order.id)
    worksheet.cell(row=row_num, column=2, value=order.product.name)
    worksheet.cell(row=row_num, column=3, value=order.order_quantity)
    worksheet.cell(row=row_num, column=4, value=order.product.price)
    worksheet.cell(row=row_num, column=5, value=order.created_by.username)
    worksheet.cell(row=row_num, column=6, value=order.date.strftime('%Y-%m-%d %H:%M:%S'))

    # Create an HTTP response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=receipt_{order.id}.xlsx'

    # Save the workbook to the response
    workbook.save(response)

    return response

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from django.shortcuts import get_object_or_404, render
from django.http import HttpResponse

def export_receiptALL(request):
    # Retrieve all orders
    orders = Order.objects.all()

    # Create a new Excel workbook and select the active worksheet
    workbook = Workbook()
    worksheet = workbook.active

    # Set column widths for the first 6 columns
    column_widths = [20, 30, 15, 20, 20, 10]  # Adjust the widths as per your requirements
    for col_num, width in enumerate(column_widths, 1):
        col_letter = get_column_letter(col_num)
        worksheet.column_dimensions[col_letter].width = width

    # Write the title
    title = "INVENTORY MANAGEMENT SYSTEM"
    title_cell = worksheet.cell(row=1, column=1, value=title)
    title_cell.alignment = Alignment(horizontal='center')
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    # Set column headers
    headers = ['Order ID', 'Product Name', 'Quantity', 'Price', 'Created By', 'Date']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = worksheet.cell(row=2, column=col_num, value=header)
        cell.alignment = Alignment(horizontal='center')

    # Write the orders data to the worksheet
    row_num = 3
    for order in orders:
        worksheet.cell(row=row_num, column=1, value=order.id)
        worksheet.cell(row=row_num, column=2, value=order.product.name)
        worksheet.cell(row=row_num, column=3, value=order.order_quantity)
        worksheet.cell(row=row_num, column=4, value=order.product.price)
        worksheet.cell(row=row_num, column=5, value=order.created_by.username)
        worksheet.cell(row=row_num, column=6, value=order.date.strftime('%Y-%m-%d %H:%M:%S'))
        row_num += 1

    # Create an HTTP response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=orders.xlsx'

    # Save the workbook to the response
    workbook.save(response)

    return response


def export_dispatch(request):
    # Retrieve all orders
    dispatchs = dispatched.objects.all()

    # Create a new Excel workbook and select the active worksheet
    workbook = Workbook()
    worksheet = workbook.active

    # Set column widths for the first 6 columns
    column_widths = [20, 30, 15, 20, 20, 10]  # Adjust the widths as per your requirements
    for col_num, width in enumerate(column_widths, 1):
        col_letter = get_column_letter(col_num)
        worksheet.column_dimensions[col_letter].width = width

    # Write the title
    title = "INVENTORY MANAGEMENT SYSTEM"
    title_cell = worksheet.cell(row=1, column=1, value=title)
    title_cell.alignment = Alignment(horizontal='center')
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    # Set column headers
    headers = ['dispatched ID', 'Dispatched Product', 'Quantity', 'Price', 'Created By', 'Date']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = worksheet.cell(row=2, column=col_num, value=header)
        cell.alignment = Alignment(horizontal='center')

    # Write the orders data to the worksheet
    row_num = 3
    for dispatch in dispatchs:
        worksheet.cell(row=row_num, column=1, value=dispatch.id)
        worksheet.cell(row=row_num, column=2, value=dispatch.product.name)
        worksheet.cell(row=row_num, column=3, value=dispatch.order_quantity)
        worksheet.cell(row=row_num, column=4, value=dispatch.product.price)
        worksheet.cell(row=row_num, column=5, value=dispatch.created_by.username)
        worksheet.cell(row=row_num, column=6, value=dispatch.date.strftime('%Y-%m-%d %H:%M:%S'))
        row_num += 1

    # Create an HTTP response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=dispatched.xlsx'

    # Save the workbook to the response
    workbook.save(response)

    return response

def export_products(request):
    # Retrieve all orders
    products = Product.objects.all()

    # Create a new Excel workbook and select the active worksheet
    workbook = Workbook()
    worksheet = workbook.active

    # Set column widths for the first 6 columns
    column_widths = [20, 30, 15, 20, 20, 10]  # Adjust the widths as per your requirements
    for col_num, width in enumerate(column_widths, 1):
        col_letter = get_column_letter(col_num)
        worksheet.column_dimensions[col_letter].width = width

    # Write the title
    title = "INVENTORY MANAGEMENT SYSTEM"
    title_cell = worksheet.cell(row=1, column=1, value=title)
    title_cell.alignment = Alignment(horizontal='center')
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    # Set column headers
    headers = ['Product ID', 'Product Name', 'Category', 'Quantity', 'Description', 'Price']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = worksheet.cell(row=2, column=col_num, value=header)
        cell.alignment = Alignment(horizontal='center')

    # Write the orders data to the worksheet
    row_num = 3
    for product in products:
        worksheet.cell(row=row_num, column=1, value=product.id)
        worksheet.cell(row=row_num, column=2, value=product.name)
        worksheet.cell(row=row_num, column=3, value=product.category)
        worksheet.cell(row=row_num, column=4, value=product.quantity)
        worksheet.cell(row=row_num, column=5, value=product.description)
        worksheet.cell(row=row_num, column=6, value=product.price)
        row_num += 1

    # Create an HTTP response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Products.xlsx'

    # Save the workbook to the response
    workbook.save(response)

    return response



import qrcode
from django.shortcuts import render, redirect
from django.http import HttpResponse
from .forms import ProductForm

def generate_qrcode(request):
    products = Product.objects.all()
    if request.method == 'POST':
        form = ProductForm(request.POST)
        if form.is_valid():
            product = form.save()
            # Generate QR code data
            data = {
                'name': product.name,
                'category': product.category,
                'quantity': product.quantity,
                'description': product.description,
                'price': product.price,
                }
            qr_data = '|'.join(f'{key}:{value}' for key, value in data.items())

            # Generate QR code image
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=10,
                border=4,
            )
            qr.add_data(qr_data)
            qr.make(fit=True)
            qr_img = qr.make_image(fill_color="black", back_color="white")

            response = HttpResponse(content_type='image/jpeg')
            qr_img.save(response, 'JPEG')
            response['Content-Disposition'] = 'attachment; filename="qrcode.jpg"'
            return response
    else:
        form = ProductForm()
    
    context = {'form': form,
               "title": "Products",
               "products": products,}
    return render(request, 'inventory/products.html', context)






from django.shortcuts import render, redirect
from .forms import ProductForm
from .models import Product

def scan_qrcodeaddproduct(request):
    if request.method == 'POST':
        form = ProductForm(request.POST)
        if form.is_valid():
            form.save()
            # Redirect to a success page or return a JSON response
            return redirect('products')
    else:
        form = ProductForm()

    context = {'form': form}
    return render(request, 'inventory/scan_qrcodeaddproduct.html', context)




# from django.shortcuts import render, redirect
# from .models import Order, dispatched

# def move_to_dispatched(request, order_id):
#     # Retrieve the specific order based on the order_id
#     order = Order.objects.get(pk=order_id)
    
#     # Create a new dispatched instance with the same data as the order
#     new_dispatched = dispatched(
#         product=order.product,
#         created_by=order.created_by,
#         order_quantity=order.order_quantity
#     )
#     new_dispatched.save()
    
#     # Delete the order from the Order class
#     order.delete()
    
#     return redirect('inventory/orders')  # Redirect to a view showing the list of orders

def addstore(request):
    stores = Store.objects.all()
    if request.method == 'POST':
        form = StoreForm(request.POST)
        if form.is_valid():
            store = form.save()

            # Generate QR code data
            data = {
                'identification': store.identification,
                'storename': store.storename,
                'address': store.address,
                'phone': store.phone,
                }
            qr_data = '|'.join(f'{key}:{value}' for key, value in data.items())

            # Generate QR code image
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=10,
                border=4,
            )
            qr.add_data(qr_data)
            qr.make(fit=True)
            qr_img = qr.make_image(fill_color="black", back_color="white")

            response = HttpResponse(content_type='image/jpeg')
            qr_img.save(response, 'JPEG')
            response['Content-Disposition'] = 'attachment; filename="qrcode.jpg"'
            return response
    else:
        form = StoreForm()
    
    context = {'form': form,
               "title": "Stores",
               "stores": stores,}
    return render(request, 'inventory/store.html', context)

def deletestore(request, pk):
    remove = Store.objects.get(id=pk)
    if request.method == 'POST':
        remove.delete()
        return redirect ('store')
    return render(request, 'inventory/deletestore.html')

def updatestore(request, pk):
    store = Store.objects.get(id=pk)
    form = StoreForm(instance=store)
    if request.method == 'POST':
        form = StoreForm(request.POST, instance=store)
        if form.is_valid():
            form.save()
            return redirect('store')

    context = {
        'form':form,
    }
    return render(request, 'inventory/updatestore.html', context)


def deleteproduct(request, pk):
    remove =Product.objects.get(id=pk)
    if request.method == 'POST':
        remove.delete()
        return redirect ('products')
    return render(request, 'inventory/deleteproduct.html')

def updateproduct(request, pk):
    product = Product.objects.get(id=pk)
    form = ProductForm(instance=product)
    if request.method == 'POST':
        form = ProductForm(request.POST, instance=product)
        if form.is_valid():
            form.save()
            return redirect('products')

    context = {
        'form':form,
    }
    return render(request, 'inventory/updateproduct.html', context)

# from django.shortcuts import render, redirect
# from .models import Order, dispatched

# def move_to_dispatched(request, order_id):
#     order = Order.objects.get(id=order_id)
    
#     if request.method == 'POST':
#         dispatched_order = dispatched.objects.create(
#             product=order.product,
#             created_by=order.created_by,
#             order_quantity=order.order_quantity
#         )
#         order.delete()
#         return redirect('orders')
    
#     context = {
#         'order': order
#     }
#     return render(request, 'inventory/dispatch.html', context)


from django.shortcuts import render, redirect
from .models import Order, dispatched
from .forms import dispatchForm

def dispatch_order(request, order_id):
    order = Order.objects.get(pk=order_id)

    if request.method == 'POST':
        form = dispatchForm(request.POST)
        if form.is_valid():
            storename = form.cleaned_data['storename']
            dispatched_order = dispatched(
                product=order.product,
                created_by=order.created_by,
                order_quantity=order.order_quantity,
                storename=storename
            )
            dispatched_order.save()
            order.delete()
            return redirect('orders')  # Replace 'orders' with the appropriate URL name for your orders list view
    else:
        form = dispatchForm()

    context = {
        'title': 'Dispatch Order',
        'form': form,
    }
    return render(request, 'inventory/dispatch.html', context)

def displaydispatch(request):
    dispatches = dispatched.objects.all()
    context = {
        'title': 'Dispatched Orders',
        'dispatches': dispatches,
    }
    return render(request, 'inventory/displaydispatch.html', context)

# from django.utils.translation import gettext_lazy
# from django.utils.translation import get_language, activate, gettext

# def translate(language):
#     cur_language = get_language()
#     try:
#         activate(language)
#         text = gettext('Products')
#     finally:
#         activate(cur_language)
#     return text


def addsupplier(request):
    suppliers = Supplier.objects.all()
    if request.method == 'POST':
        form = SupplierForm(request.POST)
        if form.is_valid():
             form.save()
    else:
        form = SupplierForm()
    
    context = {'form': form,
               "title": "Suppliers",
               "suppliers": suppliers,}
    return render(request, 'inventory/supplier.html', context)

def deletesupplier(request, pk):
    remove = Supplier.objects.get(id=pk)
    if request.method == 'POST':
        remove.delete()
        return redirect ('supplier')
    return render(request, 'inventory/deletesupplier.html')

def updatesupplier(request, pk):
    supplier = Supplier.objects.get(id=pk)
    form = SupplierForm(instance=supplier)
    if request.method == 'POST':
        form = SupplierForm(request.POST, instance=supplier)
        if form.is_valid():
            form.save()
            return redirect('supplier')

    context = {
        'form':form,
    }
    return render(request, 'inventory/updatesupplier.html', context)







from django.shortcuts import render, redirect, get_object_or_404
from django.core.mail import send_mail, EmailMessage
from django.template.loader import render_to_string
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import io
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

from .forms import OrderForm
from .models import Order, Product, Supplier

# views.py
from django.http import HttpResponse
from django.core.mail import EmailMessage
from django.shortcuts import render, redirect
from io import BytesIO
from reportlab.pdfgen import canvas
import openpyxl

# The rest of your code remains unchanged
# views.py
import openpyxl

# ... (rest of the imports and code remain the same) ...

def send_purchase_order(request):
    if request.method == 'POST':
        form = OrderForm(request.POST)
        if form.is_valid():
            order = form.save(commit=False)
            product = order.product
            order_quantity = order.order_quantity
            supplier = product.Supplier

            # Generate the Excel file and send the email
            excel_file = generate_purchase_order_excel(order, product, supplier)
            excel_bytes = excel_file.getvalue()

            subject = f'Purchase Order for {product.name}'
            message = f'Please find the attached purchase order for {product.name}.'
            from_email = 'your_email@example.com'
            to_email = [supplier.email]

            email = EmailMessage(subject, message, from_email, to_email)
            email.attach(f'Purchase_Order_{product.name}.xlsx', excel_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            email.send()
            return redirect('send_purchase_order')
        else:
            return HttpResponse("Not enough quantity available.")
    else:
        form = OrderForm()

    context = {
        "title": "Send Purchase Order",
        "form": form,
    }
    return render(request, 'inventory/send_purchase_order.html', context)

# views.py
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ... (rest of the imports and code remain the same) ...

def generate_purchase_order_excel(order, product, supplier):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Adding a header with company information
    worksheet.merge_cells('A1:E1')
    worksheet['A1'] = 'Your Company Name'
    worksheet['A1'].font = Font(size=18, bold=True)
    worksheet['A1'].alignment = Alignment(horizontal='center')
    
    # Adding purchase order title
    worksheet.merge_cells('A3:E3')
    worksheet['A3'] = 'Purchase Order'
    worksheet['A3'].font = Font(size=16, bold=True)
    worksheet['A3'].alignment = Alignment(horizontal='center')
    
    # Adding supplier information
    worksheet['A5'] = 'Supplier Information:'
    worksheet['A5'].font = Font(size=12, bold=True)
    
    worksheet['A6'] = f'Supplier: {supplier.suppliername}'
    worksheet['A7'] = f'Email: {supplier.email}'
    worksheet['A8'] = f'Address: {supplier.address}'
    worksheet['A9'] = f'Phone: {supplier.phone_number}'
    
    # Adding a table for item details
    worksheet['A11'] = 'Item'
    worksheet['B11'] = 'Description'
    worksheet['C11'] = 'Quantity'
    worksheet['D11'] = 'Price per Unit'
    worksheet['E11'] = 'Total'
    
    for col in range(1, 6):
        cell = worksheet.cell(row=11, column=col)
        cell.font = Font(size=14, bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")
        cell.border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
    # Convert order quantity and product price to numeric values
    order_quantity = int(order.order_quantity)
    product_price = float(product.price)

    # Adding the item details
    item_row = 12
    worksheet.cell(row=item_row, column=1, value=product.name)
    worksheet.cell(row=item_row, column=2, value='Item description goes here')  # Replace with actual description if available
    worksheet.cell(row=item_row, column=3, value=order.order_quantity)
    worksheet.cell(row=item_row, column=4, value=product.price)
    worksheet.cell(row=item_row, column=5, value=order_quantity * product_price)

    # Formatting the item details
    for col in range(1, 6):
        cell = worksheet.cell(row=item_row, column=col)
        cell.border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             bottom=Side(style='thin'))


    # Adding a total section
    total_row = item_row + 1

    for col in range(4, 6):
        cell = worksheet.cell(row=total_row, column=col)
        cell.font = Font(size=14, bold=True)
        cell.border = Border(top=Side(style='double'),
                             left=Side(style='thin'),
                             right=Side(style='thin'))

    worksheet.cell(row=total_row, column=4, value='Total:')
    worksheet.cell(row=total_row, column=5, value=order_quantity * product_price)


    # Adjusting column widths
    column_widths = [30, 40, 15, 20, 20]
    for i, width in enumerate(column_widths, start=1):
        worksheet.column_dimensions[get_column_letter(i)].width = width

    # Save the Excel file to a BytesIO buffer
    excel_buffer = BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)

    return excel_buffer
